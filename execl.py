from datetime import datetime
import openpyxl
import pandas as pd


def read_execle(name):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(name)
    # 选择工作表，可以通过工作表的名称或索引来选择
    sheet = workbook["Sheet1"]  # 或者通过索引：sheet = workbook.worksheets[0]
    arr = []
    # 遍历行
    for row in sheet.iter_rows(min_row=1,
                               max_row=sheet.max_row,
                               values_only=True):
        data = {"id": row[0], "name": row[1]}
        arr.append(data)
    arr.pop(0)
    return arr


def export_excel(export):
    # 将字典列表转换为DataFrame
    pf = pd.DataFrame(list(export))
    current_time = datetime.now()
    date_string = current_time.strftime("%Y-%m-%d-%H-%M")
    print(date_string)
    file_path = pd.ExcelWriter("数据/筛选" + date_string + ".xlsx")
    # 替换空单元格
    pf.fillna(" ", inplace=True)
    # 输出
    pf.to_excel(file_path, index=False)
    # 保存表格
    file_path.close()
