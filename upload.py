import openpyxl
from datetime import datetime
import copy
import pandas as pd
import requests
import json


def main():
    map = {
        "qgqp_b_id": "3305424307279569a80027fe77ecc812",
        "websitepoptg_show_time": "1700055588785",
        "mtp": "1",
        "sid": "166380500",
        "st_si": "25990893322027",
        "p_origin": "https://passport2.eastmoney.com",
        "ct":
            "HnFLR8UpntU4SKiijzfCIHXcMRF4r3NxBOirKXY8Xd7xiq0p5mJzmCzcpLul6USGh0s61630Ld4kj9II66XiBnYnmiMa0oINPfdMddqZ-th8xCqkdNK5FRoiesC11PmkMWIX97LdeOmqL5kjXIxRQLREekA_z-ch6Gi9NHFwSIU",
        "ut":
            "FobyicMgeV7bfas_M05TDDnCyhYAqcyBpyQa_M_ub0mXJre23hAnMCiWW1ndWVMwN7qNzQJU71izZri-0lJkq6bqZrWv77Z_7Wy2bkPOHBViEznOiTNQMTOGO8pDV1TyqrbohFFtCvgj_OakJYiSyNNPdHl00zKpI_uvAHbxGFOEd6EJV-l1jXYB517ZUAdzPyNHM0qteOmKX0c8Ttjs-Y7LmEnNqzhlyoVZfSKR-TTxNUKZxpMi6OWQREfOTthOvFh9T_f20GYeYILZMeQZhaNHH8IOpS7Y",
        "pi":
            "1519356260729744%3Br1519356260729744%3B%E5%BC%A0%E8%B5%B7%E7%81%B5kylin%3Bwy%2FKEu9fqBACwEI6tEdJzN2TJ2Bb7yRv6nd01CVsEj8IoO6YchOpi2vh1GfP6TEven0iYWDgr8CEW1eJSo86YZcFsfRxlwFHxGyDE2E7TGm0jJsq%2FEc%2Bz3uUxuZph6GXrIg3SGGuMEnIaaV3yWcxZFwpHZNUwyym6bc3%2BYw4XNMQZN6J9mgcReHktZf%2BeYzIRPT6dU4G%3Bcl4jlMLE19nMmKAR4ZU5in9MmM88msYf%2FKP%2FDiAtqv3ef2m3n714f4u8Yxo6CoNBZx2RvDVdZdxBPL7kzhmn%2FgUfORbMRVTZNrEVE8AVXOmWdVgTAO9xdGy%2FUxyIr5q%2FoiW9i0hbMWfys6UvJLAFzmj7DMwtew%3D%3D",
        "uidal": "1519356260729744%e5%bc%a0%e8%b5%b7%e7%81%b5kylin",
        "vtpst": "|",
        "st_asi": "delete",
        "HAList":
            "ty-0-300315-%u638C%u8DA3%u79D1%u6280%2Cty-90-BK1046-%u6E38%u620F%2Cty-90-BK1015-%u80FD%u6E90%u91D1%u5C5E%2Cty-90-BK0732-%u8D35%u91D1%u5C5E%2Cty-0-001337-%u56DB%u5DDD%u9EC4%u91D1%2Cty-1-600600-%u9752%u5C9B%u5564%u9152%2Cty-90-BK0458-%u4EEA%u5668%u4EEA%u8868%2Cty-0-000001-%u5E73%u5B89%u94F6%u884C%2Cty-1-603001-ST%u5965%u5EB7%2Cty-1-603602-%u7EB5%u6A2A%u901A%u4FE1",
        "st_pvi": "28103022399416",
        "st_sp": "2023-10-06%2019%3A33%3A20",
        "st_inirUrl": "https%3A%2F%2Fwww.google.com%2F",
        "st_sn": "12",
        "st_psi": "20231206220529104-113200301712-9860214582"
    }
    current_time = datetime.now()
    name = current_time.strftime("%Y%m%d")
    cookie = dict_to_string(map)
    gid = creat_choose_group(cookie=cookie, name=name)
    arr = read_data()
    print(arr)

    for id in arr:
        upload(cookie=cookie, id=id, gid=gid)


def dict_to_string(data):
    return "; ".join([f"{key}={value}" for key, value in data.items()])


def creat_choose_group(cookie, name):
    headers = {
        'Referer': 'http://quote.eastmoney.com/zixuan/?from=home',
        'Cookie': cookie
    }
    params = {
        "appkey": "d41d8cd98f00b204e9800998ecf8427e",
        "gn": name,
        "_": "1701871529068"
    }
    url = "http://myfavor.eastmoney.com/v4/webouter/ag"
    data = requests.get(url, params=params, headers=headers)
    json_str = json.dumps(data.text)
    res = json.loads(data.text)
    try:
        return res["data"]["gid"]
    except Exception as e:
        print("请求出错")
    print(json_str)


def read_data():
    name = "数据/筛选2023-12-26-22-23.xlsx"

    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(name)
    # 选择工作表，可以通过工作表的名称或索引来选择
    sheet = workbook["Sheet1"]  # 或者通过索引：sheet = workbook.worksheets[0]
    arr = []
    # 遍历行
    for row in sheet.iter_rows(min_row=1,
                               max_row=sheet.max_row,
                               values_only=True):
        arr.append(row[4])
    arr.pop(0)
    return arr


def upload(cookie, id, gid):
    sc = "0$"
    if id.startswith("0") == True:
        sc = "0$"
    elif id.startswith("6") == True:
        sc = "1$"
    headers = {
        'Referer': 'http://quote.eastmoney.com/zixuan/?from=home',
        'Cookie': cookie
    }
    params = {
        "appkey": "d41d8cd98f00b204e9800998ecf8427e",
        "g": gid,
        "sc": sc + id,
    }
    url = "http://myfavor.eastmoney.com/v4/webouter/as"
    data = requests.get(url, params=params, headers=headers)
    print(data)


if __name__ == "__main__":
    main()
